import os
from datetime import datetime

import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font

MESES = [
    'enero',
    'febrero',
    'marzo',
    'abril',
    'mayo',
    'junio',
    'julio',
    'agosto',
    'septiembre',
    'octubre',
    'noviembre',
    'diciembre'
]

date = datetime.now()
mes = MESES[date.month - 1]
anio = date.year

# CONCATENACIONES FECHA
anio_at_string = "AÑO" + ' ' + str(anio)
mes_at_string = f"{date.month:02d}" + " " + "-" + " "f"{mes.capitalize()}"


# C:\Users\mduran.bc\Desktop\03 - Marzo
directorio_local_diario = os.path.join("C:\\", "Users", "mduran.bc", "Desktop", "KCRM")


# C:\Users\34633\Desktop\DEMO ROBOT_AGBAR
# directorio_local_diario = os.path.join("C:\\", "Users", "34633", "Desktop", "DEMO ROBOT_AGBAR")


def contar_cuelgue_y_repeticione(ruta):
    # Leer el archivo Excel que contiene los datos
    for file_name in os.listdir(ruta):
        data = []
        num_cuelgues_por_agente = {}
        num_repeticiones_por_agente = {}
        if file_name.endswith(".xlsx"):
            file_path = os.path.join(ruta, file_name)
            print(file_path)
            data = pd.read_excel(file_path)

            for index, row in data.iterrows():
                agente = row[3]
                extension = row[19]
                print("Esto es el agente = " + str(agente))
                print("Esto es la extensión = " + str(extension) + " " + str(len(str(extension))))

                if 3 < len(str(extension)) <= 7:
                    if agente in num_cuelgues_por_agente:
                        num_cuelgues_por_agente[agente] += 1
                    else:
                        num_cuelgues_por_agente[agente] = 1

                    if agente in num_repeticiones_por_agente:
                        if extension in num_repeticiones_por_agente[agente]:
                            num_repeticiones_por_agente[agente][extension] += 1
                        else:
                            num_repeticiones_por_agente[agente][extension] = 1
                    else:
                        num_repeticiones_por_agente[agente] = {extension: 1}
        # Crear un diccionario para llevar un registro del número de veces que cada agente ha colgado la llamada

        # Iterar a través de todas las filas en las columnas 5 y 21

        print(file_path)
        # Abrir el archivo Excel
        workbook = load_workbook(filename=file_path)

        # Crear una nueva hoja llamada 'conteo_total_llamadas'
        sheet = workbook.create_sheet('CONTEO_TOTAL_LLAMADAS')

        borde = Border(left=Side(border_style='thin', color='000000'),
                       right=Side(border_style='thin', color='000000'),
                       top=Side(border_style='thin', color='000000'),
                       bottom=Side(border_style='thin', color='000000'))

        relleno_filas_vacias = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        relleno_id_agente = PatternFill(start_color='FFFCCC', end_color='FFFCCC', fill_type='solid')
        relleno_id_agente_rojo = PatternFill(start_color='CF7C89', end_color='CF7C89', fill_type='solid')

        # Escribir los encabezados de la tabla en la hoja
        sheet.cell(row=1, column=1, value='ID del agente').border = borde
        sheet.cell(row=1, column=2, value='Extensión').border = borde
        sheet.cell(row=1, column=3, value='Número de repeticiones por extensión').border = borde
        sheet.cell(row=1, column=4, value='Número de llamadas colgadas por agente').border = borde

        sheet.cell(row=1, column=1, value='ID del agente').font = Font(bold=True)
        sheet.cell(row=1, column=2, value='Extensión').font = Font(bold=True)
        sheet.cell(row=1, column=3, value='Número de repeticiones por extensión').font = Font(bold=True)
        sheet.cell(row=1, column=4, value='Número de llamadas colgadas por agente').font = Font(bold=True)

        sheet.cell(row=1, column=1, value='ID del agente').fill = relleno_filas_vacias
        sheet.cell(row=1, column=2, value='Extensión').fill = relleno_filas_vacias
        sheet.cell(row=1, column=3, value='Número de repeticiones por extensión').fill = relleno_filas_vacias
        sheet.cell(row=1, column=4, value='Número de llamadas colgadas por agente').fill = relleno_filas_vacias

        # Escribir los datos en la hoja

        row_num = 2
        for agente, num_cuelgues in num_cuelgues_por_agente.items():
            for extension, num_repeticiones in num_repeticiones_por_agente[agente].items():
                sheet.cell(row=row_num, column=1, value=agente).fill = relleno_id_agente
                sheet.cell(row=row_num, column=2, value=extension)
                sheet.cell(row=row_num, column=3, value=num_repeticiones)
                sheet.cell(row=row_num, column=4, value=num_cuelgues)
                if num_cuelgues > 3:
                    sheet.cell(row=row_num, column=1, value=agente).fill = relleno_id_agente_rojo
                # Añadimos una fila vacía y la pintamos con el color definido
                row_num += 1

            sheet.cell(row=row_num, column=1).fill = relleno_filas_vacias
            sheet.cell(row=row_num, column=2).fill = relleno_filas_vacias
            sheet.cell(row=row_num, column=3).fill = relleno_filas_vacias
            sheet.cell(row=row_num, column=4).fill = relleno_filas_vacias
            # Movemos el cursor a la siguiente fila
            row_num += 1

        # Guardar los cambios en el archivo Excel
        workbook.save(filename=file_path)
        print("Hecho")


contar_cuelgue_y_repeticione(directorio_local_diario)
