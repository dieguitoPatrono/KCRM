from collections import Counter
from datetime import datetime
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image

import pandas as pd
import os
import xlsxwriter
from openpyxl.styles import PatternFill, Border, Side, Font

MESES = [
    'enero',
    'febrero',
    'marzo',
    'abril',
    'mayo',
    'junio',
    'julio',
    'agosto',
    'septiembre',
    'octubre',
    'noviembre',
    'diciembre'
]

date = datetime.now()
mes = MESES[date.month - 1]
anio = date.year
# CONCATENACIONES FECHA
anio_at_string = "AÑO" + ' ' + str(anio)
mes_at_string = f"{date.month:02d}" + " " + "-" + " "f"{mes.capitalize()}"
# mes_at_string = f"{date.month:02d}-{mes.upper()}"

# SUBCARPETAS URL CORRECTA
# servicio = "07- AGBAR-SUEZ"
# subcarpeta1 = "01 - Listados diarios"
# subcarpeta2 = "03 - KCRM"
# subcarpeta3 = "15 - Control Interno Agentes Aguas de Barcelona"
# DIRECTORIO_RAIZ = os.path.join("U:\\", anio_at_string, servicio, subcarpeta1, subcarpeta2, subcarpeta3, mes_at_string)
# directorio_local = os.path.join("C:\\", "Users", "mduran.bc", "Desktop", "KCRM")
# directorio_local = os.path.join("C:\\", "Users", "34633", "Desktop", "DEMO ROBOT_AGBAR")

# CARPETAS
carpeta_robot = "ROBOT"


# DIRECTORIOS ESPECIFICOS
# directorio_robot = os.path.join("C:\\Users\\mduran.bc\\Desktop\\KCRM", carpeta_robot)


# CREAMOS DIRECTORIO DONDE MOVER LOS ARCHIVOS QUE YA HA UNIDO
def create_directory(folder_path):
    try:
        os.mkdir(folder_path)
    except FileExistsError:
        print("La carpeta ya existe.")
    except Exception as e:
        print("Ha ocurrido un error inesperado:", e)
    else:
        print("La carpeta se ha creado con éxito.")


# UNIFICAMOS LOS ARCHIVOS XSLM A XLSX
def merge_xslx_to_xlsx(folder_path):
    print(folder_path)
    # lista vacia
    data = []
    # por cada archivo dentro del directorio
    for file_name in os.listdir(folder_path):
        # si el archivo acaba en xlsm
        if file_name.endswith(".xlsx"):
            # establece en el directorio el archivo a tratar
            file_path = os.path.join(folder_path, file_name)
            # lo lee
            df = pd.read_excel(file_path)
            # lo añade en la lista
            data.append(df)
            # movemos el archivo que ya ha leído a una carpeta nueva llamada 'Ficheros_ya_agrupados'
            # src_file = os.path.join(directorio_local, file_name)
            # dst_folder = os.path.join(directorio_agrupados)

            # try:
            #    shutil.move(src_file, dst_folder)
            # except FileNotFoundError:
            #    print("El archivo o la carpeta de origen no existe.")
            # except Exception as e:
            #    print("Ha ocurrido un error inesperado:", e)
            # else:
            #    print("El archivo se ha movido con éxito.")
    # une el contenido leído
    merged_df = pd.concat(data)
    # nos crea un archivo llamado FCR y lo escribe
    merged_df.to_excel(
        os.path.join(folder_path, "ROBOT", "INF_CONTROL_INTERNO_SERVICIO_AGUAS_BCN_" + mes +
                     ".xlsx"),
        engine='xlsxwriter')

    # CREAMOS LA NUEVA HOJA CON LOS DATOS A TRATAR PARA EL FCR.
    # CONTEA CUANTAS VECES HA LLAMADO ESE CONTACTO.
    # DE LOS CAMPOS DE TELEFONO REPETIDOS SÓLO NOS QUEDAMOS EL PRIMERO QUE ENCUENTRA Y LOS OTROS LOS ELIMINA
    # SIN MODIFICAR EL CONTEO.
    def contar_cuelgues_y_repeticiones(ruta_del_archivo):
        # Leer el archivo Excel que contiene los datos

        data = pd.read_excel(ruta_del_archivo)

        # Crear un diccionario para llevar un registro del número de veces que cada agente ha colgado la llamada
        num_cuelgues_por_agente = {}
        num_repeticiones_por_agente = {}

        # Iterar a través de todas las filas en las columnas 5 y 21
        for index, row in data.iterrows():
            agente = row[4]
            extension = row[20]
            print("Esto es el agente = " + str(agente))
            print("Esto es la extensión = " + str(extension) + str(len(str(extension))))
            if len(str(extension)) == 7:
                if agente in num_cuelgues_por_agente:
                    num_cuelgues_por_agente[agente] += 1
                else:
                    num_cuelgues_por_agente[agente] = 1

                if agente in num_repeticiones_por_agente:
                    if extension in num_repeticiones_por_agente[agente]:
                        num_repeticiones_por_agente[agente][extension] += 1
                    else:
                        num_repeticiones_por_agente[agente][extension] = 1
                else:
                    num_repeticiones_por_agente[agente] = {extension: 1}
        print(ruta_del_archivo)
        # Abrir el archivo Excel
        workbook = load_workbook(filename=ruta_del_archivo)

        # Crear una nueva hoja llamada 'conteo_total_llamadas'
        sheet = workbook.create_sheet('CONTEO_TOTAL_LLAMADAS')

        borde = Border(left=Side(border_style='thin', color='000000'),
                       right=Side(border_style='thin', color='000000'),
                       top=Side(border_style='thin', color='000000'),
                       bottom=Side(border_style='thin', color='000000'))

        relleno_filas_vacias = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        relleno_id_agente = PatternFill(start_color='FFFCCC', end_color='FFFCCC', fill_type='solid')
        relleno_id_agente_rojo = PatternFill(start_color='CF7C89', end_color='CF7C89', fill_type='solid')

        # Escribir los encabezados de la tabla en la hoja
        # Escribir los encabezados de la tabla en la hoja
        sheet.cell(row=1, column=1, value='ID del agente').border = borde
        sheet.cell(row=1, column=2, value='Extensión').border = borde
        sheet.cell(row=1, column=3, value='Número de repeticiones por extensión').border = borde
        sheet.cell(row=1, column=4, value='Número de llamadas colgadas por agente').border = borde

        sheet.cell(row=1, column=1, value='ID del agente').font = Font(bold=True)
        sheet.cell(row=1, column=2, value='Extensión').font = Font(bold=True)
        sheet.cell(row=1, column=3, value='Número de repeticiones por extensión').font = Font(bold=True)
        sheet.cell(row=1, column=4, value='Número de llamadas colgadas por agente').font = Font(bold=True)

        sheet.cell(row=1, column=1, value='ID del agente').fill = relleno_filas_vacias
        sheet.cell(row=1, column=2, value='Extensión').fill = relleno_filas_vacias
        sheet.cell(row=1, column=3, value='Número de repeticiones por extensión').fill = relleno_filas_vacias
        sheet.cell(row=1, column=4, value='Número de llamadas colgadas por agente').fill = relleno_filas_vacias

        # Escribir los datos en la hoja
        row_num = 2
        for agente, num_cuelgues in num_cuelgues_por_agente.items():
            for extension, num_repeticiones in num_repeticiones_por_agente[agente].items():
                sheet.cell(row=row_num, column=1, value=agente).fill = relleno_id_agente
                sheet.cell(row=row_num, column=2, value=extension)
                sheet.cell(row=row_num, column=3, value=num_repeticiones)
                sheet.cell(row=row_num, column=4, value=num_cuelgues)
                if num_cuelgues > 8:
                    sheet.cell(row=row_num, column=1, value=agente).fill = relleno_id_agente_rojo
                row_num += 1
            sheet.cell(row=row_num, column=1).fill = relleno_filas_vacias
            sheet.cell(row=row_num, column=2).fill = relleno_filas_vacias
            sheet.cell(row=row_num, column=3).fill = relleno_filas_vacias
            sheet.cell(row=row_num, column=4).fill = relleno_filas_vacias
            # Movemos el cursor a la siguiente fila
            row_num += 1

        # Guardar los cambios en el archivo Excel
        workbook.save(filename=ruta_del_archivo)
        print("Hecho")

    # create_directory(os.path.join("C:\\Users\\ddela\\Desktop\\KCRM", carpeta_robot))

    merge_xslx_to_xlsx("C:\\Users\\ddela\\Desktop\\KCRM")

    contar_cuelgues_y_repeticiones(os.path.join("C:\\Users\\ddela\\Desktop\\KCRM\\ROBOT",
                                                "INF_CONTROL_INTERNO_SERVICIO_AGUAS_BCN_" + mes + ".xlsx"))
